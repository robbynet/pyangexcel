"""Excel output plugin
Generates a Excel file that presents a schema data.
"""

import optparse
import datetime
import pprint
import string
from openpyxl import Workbook
from openpyxl.styles import colors, Font, Fill, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.comments import Comment

from pyang import plugin
from pyang import statements
from pyang import util

headlineStyle = NamedStyle(name="headlineStyle")
headlineStyle.font = Font(bold=True, size=14)
# bd = Side(style='thick', color="000000")
# headlineStyle.border = Border(bottom=Side(style='double'))

columnStyle = NamedStyle(name="columnStyle")
columnStyle.font = Font(bold=True, size=10)
bd = Side(style='thick', color="000000")
columnStyle.border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))
columnStyle.fill = PatternFill(start_color='00C0C0C0',
                   end_color='00C0C0C0',
                   fill_type='solid')

infoStyle = NamedStyle(name="infoStyle")
infoStyle.font = Font(italic=True, size=10)
infoStyle.alignment = Alignment(horizontal='left', vertical='top', wrapText=False)
# infoStyle.alignment.wrap_text=False

infoBoldStyle = NamedStyle(name="infoBoldStyle")
infoBoldStyle.font = Font(italic=True, size=10, bold=True)
infoBoldStyle.alignment = Alignment(horizontal='left', vertical='top', wrapText=False)
# infoBoldStyle.alignment.wrap_text=False

normalStyle = NamedStyle(name="normalStyle")
normalStyle.font = Font(size=10)
# normalStyle.alignment.wrap_text=False
normalStyle.border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))

skipStyle = NamedStyle(name="skipStyle")
skipStyle.font = Font(size=10)
# skipStyle.alignment.wrap_text=False
skipStyle.border = Border(top=Side(style='thin'),
                bottom=Side(style='thin'))

class excel:
    def __init__(self, name="sample"):
        self.name = name
        self.wb = Workbook()
        self.start_col = 1
        self.start_row = 1
        self.wb.add_named_style(headlineStyle)
        self.wb.add_named_style(columnStyle)
        self.wb.add_named_style(infoStyle)
        self.wb.add_named_style(infoBoldStyle)
        self.wb.add_named_style(normalStyle)
        self.wb.add_named_style(skipStyle)

    def sheet(self, sheetname="sample", sheettitle="Title of Sheet", start_col=2, start_row=3, titlelen=1):
        if hasattr(self, 'ws'):
            self.ws = self.wb.create_sheet(title=sheetname)
        else:
            self.ws = self.wb.active
            self.ws.title = sheetname
        self.start_col = start_col
        self.start_row = start_row
        self.ws.row_dimensions[start_row].height = 40
        c = self.ws.cell(column=start_col, row=start_row, value=sheettitle)
        c.style = 'Title'
        # c.alignment = Alignment(horizontal="center", vertical="center")
        if titlelen > 1:
            self.ws.merge_cells(start_row=start_row, start_column=start_col,
                    end_row=start_row, end_column=start_col+(titlelen - 1))


    def save(self):
        if self.wb:
            self.wb.save(self.name + ".xlsx")

    def write(self, title="", info="", data=[[]], col=0, row=0, titlelen=1, info_height=40):
        ws = self.ws
        if row <= 0:
            if ws._current_row:
                row = ws._current_row + 2
            else:
                row = self.start_row
        if col <= 0:
            col = self.start_col
        # print col, row
        if title:
            cur_cell = ws.cell(column=col, row=row, value=title)
            if titlelen > 1:
                ws.merge_cells(start_row=row, start_column=col,
                        end_row=row, end_column=col+(titlelen - 1))
            cur_cell.style = 'headlineStyle'
            row += 1
        if info:
            if isinstance(info, dict):
                for key, value in info.items():
                    cur_cell = ws.cell(column=col, row=row, value=key)
                    cur_cell.style = 'infoBoldStyle'
                    lines = value.splitlines()
                    for offset, line in enumerate(lines):
                        cur_cell = ws.cell(column=col, row=row + offset + 1, value=line)
                        cur_cell.style = 'infoStyle'
                    row += len(lines) + 2
            else:
                lines = info.splitlines()
                for offset, line in enumerate(lines):
                    cur_cell = ws.cell(column=col, row=row+offset, value=line)
                    cur_cell.style = 'infoStyle'
                row += len(lines)
            row += 1
        if data and isinstance(data, (list, tuple)) and data[0]:
            cur_cell = ws.cell(column=col, row=row, value="No")
            cur_cell.style = 'columnStyle'
            offset = 0
            for j, d in enumerate(data[0]):
                if isinstance(d, dict):
                    cur_cell = ws.cell(column=col+j+1+offset, row=row, value=d['value'])
                    if 'merge' in d and d['merge'] > 0:
                        ws.merge_cells(start_row=row, start_column=col+j+1+offset,
                            end_row=row, end_column=col+j+1+offset+d['merge'])
                        offset += d['merge']
                else:
                    cur_cell = ws.cell(column=col+j+1+offset, row=row, value=d)
                cur_cell.style = 'columnStyle'
            row += 1
            for i, r in enumerate(data[1:]):
                offset = 0
                cur_cell = ws.cell(row=row+i, column=col, value=i+1)
                cur_cell.style = 'normalStyle'
                for j, d in enumerate(r):
                    # print(i, j, d)
                    if isinstance(d, dict):
                        if 'max_skip' in d:
                            for skip_idx in range(d['max_skip'] + 1):
                                cur_cell = ws.cell(row=row+i, column=col+offset+j+1+skip_idx)
                                if 'skip' in d and d['skip'] <= skip_idx:
                                    cur_cell.style = 'skipStyle'
                                else:
                                    cur_cell.style = 'normalStyle'
                        if 'skip' in d:
                            offset += d['skip']
                        cur_cell = ws.cell(row=row+i, column=col+offset+j+1, value=d['value'])
                        if 'skip' not in d:
                            cur_cell.style = 'normalStyle'
                        if 'color' in d:
                            cur_cell.fill = PatternFill(bgColor=d['color'], fill_type="solid")
                        if 'tooltip' in d and d['tooltip']:
                            cur_cell.comment = Comment(d['tooltip'], 'pyang-excel', height=400, width=500)
                        if 'max_skip' in d:
                            offset += (d['max_skip'] - d['skip'])
                    else:
                        cur_cell = ws.cell(row=row+i, column=col+offset+j+1, value=d)
                        cur_cell.style = 'normalStyle'
        elif data and isinstance(data, dict):
            for k, v in data.items():
                cur_cell = ws.cell(column=col, row=row, value=k)
                cur_cell.style = 'columnStyle'
                cur_cell = ws.cell(column=col+1, row=row, value=v)
                cur_cell.style = 'normalStyle'
                row += 1
        # print (ws._current_row)

    def adjust(self, cellwidth=None):
        dims = {}
        for row in self.ws.rows:
            for c, cell in enumerate(row):
                if cell.style != 'normalStyle' and cell.style != 'columnStyle' and cell.style != 'skipStyle':
                    continue
                try:
                    if cellwidth[c]:
                        # print(cell.value, cellwidth[c])
                        if isinstance(cellwidth[c], float):
                            cellsize = max((dims.get(cell.column_letter, 0), len(str(cell.value))*cellwidth[c]))
                            dims[cell.column_letter] = cellsize
                        else:
                            dims[cell.column_letter] = cellwidth[c]
                        continue
                except (IndexError, TypeError):
                    pass
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            # print (col, value)
            self.ws.column_dimensions[col].width = value * (12.0/16.0) + 1


helpstring="""[Terms]

  <flags> is one of:
    rw  for configuration data
    ro  for non-configuration data, output parameters to rpcs
        and actions, and notification parameters
    -w  for input parameters to rpcs and actions
    -x  for rpcs and actions
    -n  for notifications

   If the node is augmented into the tree from another module, its
   name is printed as <prefix>:<name>.

  <opts> is one of:
    ?  for an optional leaf, choice, anydata or anyxml
    !  for a presence container
    *  for a leaf-list or list
    [<keys>] for a list's keys
"""

def pyang_plugin_init():
    plugin.register_plugin(ExcelPlugin())

class ExcelPlugin(plugin.PyangPlugin):
    def add_output_format(self, fmts):
        self.multiple_modules = True
        fmts['excel'] = self

    def add_opts(self, optparser):
        optlist = [
            optparse.make_option("--excel-no-path",
                                 dest="excel_no_path",
                                 action="store_true",
                                 help="""Do not include paths to make
                                       page less wide"""),
            optparse.make_option("--excel-path",
                                 dest="excel_path",
                                 help="Subtree to print"),
            optparse.make_option("--excel-model-name",
                                 dest="excel_model_name",
                                 help="The model name of the excel"),
            optparse.make_option("--excel-model-ver",
                                 dest="excel_model_ver",
                                 help="The model version of the excel"),
            optparse.make_option("--excel-model-author",
                                 dest="excel_model_author",
                                 help="The model author of the excel"),
            ]

        g = optparser.add_option_group("Excel output specific options")
        g.add_options(optlist)

    def setup_fmt(self, ctx):
        ctx.implicit_errors = False

    def emit(self, ctx, modules, fd):
        if ctx.opts.excel_path is not None:
            path = ctx.opts.excel_path.split('/')
            if path[0] == '':
                path = path[1:]
        else:
            path = None
        modelname = ctx.opts.excel_model_name
        ex = excel("%s YANG data model specification" %modelname)
        ex.sheet(sheetname="Brief", sheettitle="%s YANG data model"%modelname)
        data = {
            "Model": modelname,
            "Version": ctx.opts.excel_model_ver,
            "Author": ctx.opts.excel_model_author,
        }
        ex.write(info="""This document is automatically generated by pyang excel plugin
in order to represent %s YANG Data Model via readable excel tables.
The %s YANG Data Model consists of a number of YANG data modules listed below ."""
%(modelname, modelname), data=data)

        ex.write(info=helpstring, info_height=200)

        data = get_namespace(modules)
        ex.write(title="%s YANG Data Module list" %modelname, data=data)
        ex.adjust()

        for module in modules:
            data, max_depth = get_module_data(module, fd, ctx, path)
            # pprint.pprint(data)
            ex.sheet(sheetname=module.arg, sheettitle="%s YANG Data Module"%module.arg)
            info = dict()
            # info_list = ['description', 'organization', 'contact', 'revision']
            info_list = ['description', 'organization', 'revision']
            for entry in info_list:
                e = module.search_one(entry)
                if e is not None:
                    istr = str(e.arg).strip()
                    if istr:
                        info[entry] = istr
                    # info = info + "[%s]\n %s\n\n" %(entry, e.arg)
            # info = info.rstrip()
            ex.write(info=info)

            if max_depth > 0:
                cellwidth = [10, 0] + [0.3]*(max_depth)
            else:
                cellwidth = [10, 0]
            ex.write(data=data, titlelen=len(data[0]) + max_depth)
            ex.adjust(cellwidth=cellwidth)
        ex.save()


def get_namespace(modules):
    module_list = [["Type", "Name", "belongs-to", "Namespace", "Prefix"]]
    for module in modules:
        bstr = ""
        b = module.search_one('belongs-to')
        if b is not None:
            bstr = "%s" % b.arg

        nsstr = ""
        ns = module.search_one('namespace')
        if ns is not None:
            nsstr = ns.arg
        pr = module.search_one('prefix')

        prstr = ""
        if pr is not None:
            prstr = pr.arg

        if module.keyword == 'module':
            module_list.append([str(module.keyword.capitalize()), module.arg, bstr, nsstr, prstr])
            # print("%s, %s%s, %s, %s"% (module.keyword.capitalize(),
            #             module.arg,
            #             bstr,
            #             nsstr,
            #             prstr))
        else:
            module_list.append([str(module.keyword.capitalize()), module.arg, bstr, "", ""])
            # print("%s, %s, %s" % (module.keyword.capitalize(), module.arg, bstr))
    return module_list


levelcnt = [0]*100
depth = 0

def get_module_data(module, fd, ctx, path):
    global levelcnt
    global depth
    max_depth = 0
    exdata = [[{"value": "Name", "merge": 0}, "Schema", "Type", "Flags", "Opts", "Schema Path"]]
    bstr = ""
    b = module.search_one('belongs-to')
    if b is not None:
        bstr = " (belongs-to %s)" % b.arg
    ns = module.search_one('namespace')
    if ns is not None:
        nsstr = ns.arg
    pr = module.search_one('prefix')
    if pr is not None:
        prstr = pr.arg
    else:
        prstr = ""

    levelcnt[1] += 1
    chs = [ch for ch in module.i_children
            if ch.keyword in statements.data_definition_keywords]
    if path is not None and len(path) > 0:
        chs = [ch for ch in chs if ch.arg == path[0]]
        path = path[1:]

    if len(chs) > 0:
        depth = 0
        print_children(chs, module, fd, ' ', path, ctx, 0, exdata=exdata)
        max_depth = max(max_depth, depth)

    for augment in module.search('augment'):
        if (hasattr(augment, 'i_target_node') and
            hasattr(augment.i_target_node, 'i_module')):
            # print(augment, augment.keyword, augment.arg)
            desc = augment.search_one('description')
            if not desc:
                desc = ""
            data = [{"skip": 0, "value":str(augment.arg), "tooltip": str(desc)}, str(augment.keyword), "-", "", "", str(augment.arg)]
            exdata.append(data)
            depth = 0
            print_children(augment.i_children, module, fd, '  ', path, ctx, 1, exdata=exdata)
            max_depth = max(max_depth, depth)

    rpcs = module.search('rpc')
    if path is not None:
        if len(path) > 0:
            rpcs = [rpc for rpc in rpcs if rpc.arg == path[0]]
            path = path[1:]
        else:
            rpcs = []
    # pprint.pprint(rpcs)

    levelcnt[1] += 1
    if len(rpcs) > 0:
        depth = 0
        print_children(rpcs, module, fd, ' ', path, ctx, 0, exdata=exdata)
        max_depth = max(max_depth, depth)

    notifs = module.search('notification')
    if path is not None:
        if len(path) > 0:
            notifs = [n for n in notifs if n.arg == path[0]]
            path = path[1:]
        else:
            notifs = []
    # pprint.pprint(notifs)

    levelcnt[1] += 1
    if len(notifs) > 0:
        depth = 0
        print_children(notifs, module, fd, ' ', path, ctx, 0, exdata=exdata)
        max_depth = max(max_depth, depth)
    if max_depth > 0:
        exdata[0][0]["merge"] = max_depth
        for each in exdata[1:]:
            each[0]["max_skip"] = max_depth
    return exdata, max_depth


def print_children(i_children, module, fd, prefix, path, ctx, level=0, exdata=None):
    global depth
    if len(i_children) > 0:
        depth = max(depth, level)
    for ch in i_children:
        disabled_feature = False
        iffeature = ch.search('if-feature')
        for f in iffeature:
            if f.arg not in ctx.features[module.arg]:
                print (ch.search("feature"), ch.search('if-feature'), ctx.features[module.arg], "disabled feature")
                # ignore disabled feature
                disabled_feature = True
        if not disabled_feature:
            print_node(ch, module, fd, prefix, path, ctx, level, exdata=exdata)

def print_node(s, module, fd, prefix, path, ctx, level=0, exdata=None):

    global levelcnt
    status = get_status_str(s)
    nodetype = ''
    options = ''
    folder = False
    ignored_node = False
    if s.i_module.i_modulename == module.i_modulename:
        name = s.arg
    else:
        name = s.i_module.i_prefix + ':' + s.arg

    pr = module.search_one('prefix')
    if pr is not None:
        prstr = pr.arg
    else:
        prstr = ""

    descr = s.search_one('description')
    descrstring = "No description"
    if descr is not None:
        descrstring = descr.arg
    flags = get_flags_str(s)
    if s.keyword == 'list':
        folder = True
    elif s.keyword == 'container':
        folder = True
        p = s.search_one('presence')
        if p is not None:
            pr_str = p.arg
            options = "!"
    elif s.keyword  == 'choice':
        folder = True
        m = s.search_one('mandatory')
        if m is None or m.arg == 'false':
            name = '(' + s.arg + ')'
            options = 'Choice'
        else:
            name = '(' + s.arg + ')'
        # ignored_node = True
    elif s.keyword == 'case':
        folder = True
        # fd.write(':(' + s.arg + ')')
        name = ':(' + s.arg + ')'
        # ignored_node = True
    elif s.keyword == 'input':
        folder = True
    elif s.keyword == 'output':
        folder = True
    elif s.keyword == 'rpc':
        folder = True
    elif s.keyword == 'notification':
        folder = True
    else:
        if s.keyword == 'leaf-list':
            options = '*'
        elif s.keyword == 'leaf' and not hasattr(s, 'i_is_key'):
            m = s.search_one('mandatory')
            if m is None or m.arg == 'false':
                options = '?'
        nodetype = get_typename(s)

    if s.keyword == 'list' and s.search_one('key') is not None:
        name += '[' + s.search_one('key').arg +  ']'

    descr = s.search_one('description')
    if descr is not None:
        descrstring = ''.join([x for x in descr.arg if ord(x) < 128])
    else:
        descrstring = ""
    levelcnt[level] += 1
    idstring = str(levelcnt[1])

    for i in range(2,level+1):
        idstring += '-' + str(levelcnt[i])

    pathstr = ""
    if not ctx.opts.excel_no_path:
        pathstr = statements.mk_path_str(s, True)

    keyword = s.keyword
    if not ignored_node:
        # print level
        if folder:
            # print ("print_node1: ", idstring, level, descrstring, name)
            # print ("print_node2: ", s.keyword, nodetype, flags, options, status, pathstr)
            # print (str(name), str(s.keyword), str(nodetype), str(flags), str(pathstr), str(status), str(descrstring))
            if exdata:
                data = [{"skip": level, "value":str(name), "tooltip": str(descrstring)}, str(s.keyword), "-", str(flags), str(options), str(pathstr)]
                exdata.append(data)
        else:
            if s.keyword in ['action', ('tailf-common', 'action')]:
                classstring = "action"
                typeinfo = action_params(s)
                typename = "parameters"
                keyword = "action"
            elif s.keyword == 'rpc' or s.keyword == 'notification':
                classstring = "folder"
                typeinfo = action_params(s)
                typename = "parameters"
            else:
                classstring = s.keyword
                typeinfo = typestring(s)
                typename = nodetype
            # print (str(name), str(classstring), str(typeinfo), str(flags), str(pathstr), str(status), str(descrstring))
            # print ("print_node3: ", idstring, level, classstring, descrstring, name, keyword, typeinfo, typename, flags, options, status, pathstr)
            if exdata:
                # typ = str(typeinfo).splitlines()
                # typ_tooltip = ' '.join(typ[1:])
                # typ = typ[0]
                typ = typename
                typ_tooltip = str(typeinfo)
                data = [{"skip": level, "value":str(name), "tooltip": str(descrstring)}, str(classstring), {"value": typ, "tooltip": typ_tooltip}, str(flags), str(options), str(pathstr)]
                exdata.append(data)
    if hasattr(s, 'i_children'):
        level += 1
        chs = s.i_children
        if path is not None and len(path) > 0:
            chs = [ch for ch in chs
                   if ch.arg == path[0]]
            path = path[1:]
        if s.keyword in ['choice', 'case']:
            print_children(chs, module, fd, prefix, path, ctx, level, exdata=exdata)
        else:
            print_children(chs, module, fd, prefix, path, ctx, level, exdata=exdata)

def get_status_str(s):
    status = s.search_one('status')
    if status is None or status.arg == 'current':
        return 'current'
    else:
        return status

def get_flags_str(s):
    if s.keyword == 'rpc':
        return '-x'
    elif s.keyword == 'notification':
        return ''
    elif s.keyword == 'input':
        return "-w"
    elif s.keyword == 'output':
        return ''
    elif s.i_config:
        return 'ro'
    else:
        return 'rw'

def get_typename(s):
    t = s.search_one('type')
    if t is not None:
        return t.arg
    else:
        return ''

def typestring(node):

    def get_nontypedefstring(node):
        s = ""
        found  = False
        t = node.search_one('type')
        if t is not None:
            s = t.arg + '\n'
            if t.arg == 'enumeration':
                found = True
                s = s + ' : {'
                for enums in t.substmts:
                    s = s + enums.arg + ','
                s = s + '}'
            elif t.arg == 'leafref':
                found = True
                s = s + ' : '
                p = t.search_one('path')
                if p is not None:
                    s = s + p.arg

            elif t.arg == 'identityref':
                found = True
                b = t.search_one('base')
                if b is not None:
                    s = s + ' {' + b.arg + '}'

            elif t.arg == 'union':
                found = True
                uniontypes = t.search('type')
                s = s + '{' + uniontypes[0].arg
                for uniontype in uniontypes[1:]:
                    s = s + ', ' + uniontype.arg
                s = s + '}'

            typerange = t.search_one('range')
            if typerange is not None:
                found = True
                s = s + ' [' + typerange.arg + ']'
            length = t.search_one('length')
            if length is not None:
                found = True
                s = s + ' {length = ' + length.arg + '}'

            pattern = t.search_one('pattern')
            if pattern is not None: # truncate long patterns
                found = True
                s = s + ' {pattern = ' + pattern.arg + '}'
        return s

    s = get_nontypedefstring(node)

    if s != "":
        t = node.search_one('type')
        # chase typedef
        type_namespace = None
        i_type_name = None
        prefix, name = util.split_identifier(t.arg)
        if prefix is None or t.i_module.i_prefix == prefix:
            # check local typedefs
            pmodule = node.i_module
            typedef = statements.search_typedef(t, name)
        else:
            # this is a prefixed name, check the imported modules
            err = []
            pmodule = util.prefix_to_module(t.i_module, prefix, t.pos, err)
            if pmodule is None:
                return
            typedef = statements.search_typedef(pmodule, name)
        if typedef is not None:
            s = s + get_nontypedefstring(typedef)
    return s

def action_params(action):
    s = ""
    for params in action.substmts:

        if params.keyword == 'input':
            inputs = params.search('leaf')
            inputs += params.search('leaf-list')
            inputs += params.search('list')
            inputs += params.search('container')
            inputs += params.search('anyxml')
            inputs += params.search('uses')
            for i in inputs:
                s += ' in: ' + i.arg + "\n"

        if params.keyword == 'output':
            outputs = params.search('leaf')
            outputs += params.search('leaf-list')
            outputs += params.search('list')
            outputs += params.search('container')
            outputs += params.search('anyxml')
            outputs += params.search('uses')
            for o in outputs:
                s += ' out: ' + o.arg + "\n"
    return s
